from __future__ import annotations

from io import BytesIO

from django.db import IntegrityError
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.http import FileResponse
from rest_framework import permissions, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.response import Response

from academics.models import ClassTeacher
from academics.models import Subject
from students.models import Student
from users.permissions import IsAdmin, IsStudent, IsTeacher
from users.rbac_permissions import HasPortalPermission

from exam_management.models import AuditLog, Exam, Promotion, Result, StudentExamMark
from exam_management.permissions import IsAdminOrTeacherWithClass
from exam_management.serializers import (
    AuditLogSerializer,
    ExamSerializer,
    PromotionSerializer,
    ResultSerializer,
    StudentExamMarkSerializer,
)
from exam_management.services.analytics import exam_analytics
from exam_management.services.audit import create_audit_log
from exam_management.services.exporters import excel_response, export_results_excel
from exam_management.services.pdf import build_report_card_pdf
from exam_management.services.result_engine import generate_results_for_exam, publish_results_for_exam


class ExamViewSet(viewsets.ModelViewSet):
    queryset = Exam.objects.select_related("class_name", "subject").all()
    serializer_class = ExamSerializer
    rbac_path = "/portal/exam/exams"
    rbac_action_map = {"generate_results": "edit", "publish": "edit", "analytics": "view", "export_excel": "view"}

    def get_permissions(self):
        if self.action in {"create", "update", "partial_update", "destroy", "generate_results", "publish"}:
            self.permission_classes = [permissions.IsAuthenticated, HasPortalPermission, IsAdmin]
        else:
            self.permission_classes = [permissions.IsAuthenticated, HasPortalPermission]
        return super().get_permissions()

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj.status == Exam.STATUS_PUBLISHED:
            raise ValidationError({"detail": "Cannot delete a published exam."})
        try:
            create_audit_log(action_type=AuditLog.ACTION_DELETE, user=request.user, details={"model": "Exam", "id": obj.id})
            return super().destroy(request, *args, **kwargs)
        except ProtectedError:
            raise ValidationError({"detail": "Cannot delete: exam has protected related records."})

    @action(detail=True, methods=["post"], url_path="generate-results")
    def generate_results(self, request, pk=None):
        exam = self.get_object()
        require_all = bool(request.data.get("require_all_marks", True))
        try:
            summary = generate_results_for_exam(exam=exam, actor=request.user, require_all_marks=require_all)
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response(
            {
                "exam": summary.exam_id,
                "students_processed": summary.students_processed,
                "results_generated": summary.results_generated,
                "missing_marks": summary.missing_marks,
            }
        )

    @action(detail=True, methods=["post"], url_path="publish")
    def publish(self, request, pk=None):
        exam = self.get_object()
        note = (request.data.get("note") or "").strip()
        try:
            updated = publish_results_for_exam(exam=exam, actor=request.user, note=note)
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response({"exam": exam.id, "published": True, "results_updated": updated})

    @action(detail=True, methods=["get"], url_path="analytics")
    def analytics(self, request, pk=None):
        exam = self.get_object()
        return Response(exam_analytics(exam=exam))

    @action(detail=True, methods=["get"], url_path="export-excel")
    def export_excel(self, request, pk=None):
        exam = self.get_object()
        content, filename = export_results_excel(exam=exam)
        return excel_response(content=content, filename=filename)


class StudentExamMarkViewSet(viewsets.ModelViewSet):
    queryset = StudentExamMark.objects.select_related("student", "exam", "subject", "exam__class_name").all()
    serializer_class = StudentExamMarkSerializer
    rbac_path = "/portal/exam/marks"
    rbac_action_map = {"bulk_upload": "create", "sheet": "view", "import_excel": "create", "sample_excel": "view"}

    def get_permissions(self):
        if self.action in {"create", "update", "partial_update", "destroy", "bulk_upload"}:
            self.permission_classes = [permissions.IsAuthenticated, HasPortalPermission, IsAdmin | IsTeacher]
        else:
            self.permission_classes = [permissions.IsAuthenticated, HasPortalPermission]
        return super().get_permissions()

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        exam_id = self.request.query_params.get("exam")
        student_id = self.request.query_params.get("student")
        if exam_id:
            qs = qs.filter(exam_id=exam_id)
        if student_id:
            qs = qs.filter(student_id=student_id)

        if getattr(user, "role", None) == "TEACHER":
            assignments = list(
                ClassTeacher.objects.filter(teacher__user_id=user.id).values_list("school_class_id", "section")
            )
            if not assignments:
                return qs.none()
            clause = Q()
            for class_id, section in assignments:
                clause |= Q(exam__class_name_id=class_id, exam__section=(section or ""))
            qs = qs.filter(clause)
        return qs

    def perform_create(self, serializer):
        if getattr(self.request.user, "role", None) == "TEACHER":
            exam = serializer.validated_data.get("exam")
            ok = ClassTeacher.objects.filter(
                school_class_id=exam.class_name_id, section=(exam.section or ""), teacher__user_id=self.request.user.id
            ).exists()
            if not ok:
                raise PermissionDenied("Not your assigned class.")
        obj = serializer.save()
        create_audit_log(action_type=AuditLog.ACTION_CREATE, user=self.request.user, details={"model": "StudentExamMark", "id": obj.id})

    def perform_update(self, serializer):
        if getattr(self.request.user, "role", None) == "TEACHER":
            exam = serializer.instance.exam
            ok = ClassTeacher.objects.filter(
                school_class_id=exam.class_name_id, section=(exam.section or ""), teacher__user_id=self.request.user.id
            ).exists()
            if not ok:
                raise PermissionDenied("Not your assigned class.")
        obj = serializer.save()
        create_audit_log(action_type=AuditLog.ACTION_UPDATE, user=self.request.user, details={"model": "StudentExamMark", "id": obj.id})

    def perform_destroy(self, instance):
        if instance.exam.status == Exam.STATUS_PUBLISHED:
            raise ValidationError({"detail": "Cannot modify marks after results are published."})
        create_audit_log(action_type=AuditLog.ACTION_DELETE, user=self.request.user, details={"model": "StudentExamMark", "id": instance.id})
        return super().perform_destroy(instance)

    @action(detail=False, methods=["post"], url_path="bulk-upload")
    def bulk_upload(self, request):
        """
        Expected payload:
        {
          "exam": 1,
          "items": [{"student": 10, "subject": 3, "marks_obtained": 78.5, "remarks": ""}, ...]
        }
        """
        exam_id = request.data.get("exam")
        items = request.data.get("items") or []
        if not exam_id:
            raise ValidationError({"exam": "exam is required."})
        if not isinstance(items, list):
            raise ValidationError({"items": "items must be a list."})

        try:
            exam = Exam.objects.select_related("class_name").get(id=exam_id)
        except Exam.DoesNotExist:
            raise ValidationError({"exam": "Exam not found."})
        if exam.status == Exam.STATUS_PUBLISHED:
            raise ValidationError({"detail": "Cannot modify marks after results are published."})

        if getattr(request.user, "role", None) == "TEACHER":
            ok = ClassTeacher.objects.filter(
                school_class_id=exam.class_name_id, section=(exam.section or ""), teacher__user_id=request.user.id
            ).exists()
            if not ok:
                raise PermissionDenied("Not your assigned class.")

        updated = 0
        created = 0
        errors = []
        for idx, it in enumerate(items):
            sid = it.get("student")
            subid = it.get("subject")
            marks = it.get("marks_obtained", None)
            remarks = (it.get("remarks") or "").strip()
            if not sid or not subid:
                continue
            if getattr(exam, "subject_id", None) and int(subid) != int(exam.subject_id):
                errors.append(
                    {
                        "index": idx,
                        "student": sid,
                        "subject": subid,
                        "error": f"Only subject_id={exam.subject_id} is allowed for this exam.",
                    }
                )
                continue
            try:
                obj, was_created = StudentExamMark.objects.update_or_create(
                    student_id=sid,
                    exam=exam,
                    subject_id=subid,
                    defaults={"marks_obtained": marks, "remarks": remarks},
                )
                created += 1 if was_created else 0
                updated += 0 if was_created else 1
            except (ValidationError, IntegrityError) as e:
                errors.append({"index": idx, "student": sid, "subject": subid, "error": str(e)})

        create_audit_log(
            action_type=AuditLog.ACTION_MARKS_BULK_UPLOAD,
            user=request.user,
            details={"exam": exam.id, "created": created, "updated": updated, "errors": len(errors)},
        )
        return Response({"exam": exam.id, "created": created, "updated": updated, "errors": errors})

    @action(detail=False, methods=["get"], url_path="sheet")
    def sheet(self, request):
        exam_id = request.query_params.get("exam")
        if not exam_id:
            raise ValidationError({"exam": "exam is required."})
        try:
            exam = Exam.objects.select_related("class_name").get(id=exam_id)
        except Exam.DoesNotExist:
            raise ValidationError({"exam": "Exam not found."})

        if getattr(request.user, "role", None) == "TEACHER":
            ok = ClassTeacher.objects.filter(
                school_class_id=exam.class_name_id, section=(exam.section or ""), teacher__user_id=request.user.id
            ).exists()
            if not ok:
                raise PermissionDenied("Not your assigned class.")

        subjects_qs = Subject.objects.filter(school_class_id=exam.class_name_id)
        if getattr(exam, "subject_id", None):
            subjects_qs = subjects_qs.filter(id=exam.subject_id)
        students_qs = Student.objects.filter(school_class_id=exam.class_name_id)
        if exam.class_name.sections:
            subjects_qs = subjects_qs.filter(section=(exam.section or ""))
            students_qs = students_qs.filter(section=(exam.section or ""))
        else:
            subjects_qs = subjects_qs.filter(section="")
            students_qs = students_qs.filter(section="")

        subjects = list(
            subjects_qs.order_by("code", "id").values("id", "code", "name", "full_marks", "pass_marks")
        )
        students = list(students_qs.order_by("first_name", "last_name", "id").values("id", "first_name", "last_name"))
        marks = list(
            StudentExamMark.objects.filter(exam=exam, student_id__in=[s["id"] for s in students], subject_id__in=[s["id"] for s in subjects])
            .values("id", "student_id", "subject_id", "marks_obtained", "grade", "remarks")
        )

        return Response(
            {
                "exam": {
                    "id": exam.id,
                    "exam_name": exam.exam_name,
                    "class_name": exam.class_name_id,
                    "classroom_label": exam.classroom_label,
                    "status": exam.status,
                },
                "subjects": subjects,
                "students": [{"id": s["id"], "name": f"{s['first_name']} {s['last_name']}".strip()} for s in students],
                "marks": marks,
            }
        )

    @action(detail=False, methods=["get"], url_path="sample-excel")
    def sample_excel(self, request):
        exam_id = request.query_params.get("exam")
        if not exam_id:
            raise ValidationError({"exam": "exam is required."})
        try:
            exam = Exam.objects.select_related("class_name").get(id=exam_id)
        except Exam.DoesNotExist:
            raise ValidationError({"exam": "Exam not found."})

        try:
            import openpyxl
        except Exception:  # pragma: no cover
            # Fallback: provide a CSV template which opens in Excel.
            import csv
            from io import StringIO

            subjects_qs = Subject.objects.filter(school_class_id=exam.class_name_id)
            if getattr(exam, "subject_id", None):
                subjects_qs = subjects_qs.filter(id=exam.subject_id)
            if exam.class_name.sections:
                subjects_qs = subjects_qs.filter(section=(exam.section or ""))
            else:
                subjects_qs = subjects_qs.filter(section="")

            students_qs = Student.objects.filter(school_class_id=exam.class_name_id)
            if exam.class_name.sections:
                students_qs = students_qs.filter(section=(exam.section or ""))
            else:
                students_qs = students_qs.filter(section="")

            sio = StringIO()
            writer = csv.writer(sio)
            writer.writerow(["student_id", "subject_id", "marks_obtained", "remarks"])
            writer.writerow([])
            writer.writerow(["# Reference: Subjects"])
            writer.writerow(["subject_id", "code", "name", "full_marks", "pass_marks"])
            for subj in subjects_qs.order_by("code", "id"):
                writer.writerow([subj.id, subj.code, subj.name, subj.full_marks, subj.pass_marks])
            writer.writerow([])
            writer.writerow(["# Reference: Students (first 50)"])
            writer.writerow(["student_id", "name"])
            for st in students_qs.order_by("first_name", "last_name", "id")[:50]:
                writer.writerow([st.id, f"{st.first_name} {st.last_name}".strip()])

            buff = BytesIO(sio.getvalue().encode("utf-8-sig"))
            buff.seek(0)
            filename = f"marks_upload_exam_{exam.id}.csv"
            return FileResponse(buff, as_attachment=True, filename=filename, content_type="text/csv")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MarksUpload"
        ws.append(["student_id", "subject_id", "marks_obtained", "remarks"])

        ws2 = wb.create_sheet("Reference")
        ws2.append(["Exam", exam.exam_name])
        ws2.append(["Class", exam.class_name.name if exam.class_name_id else ""])
        ws2.append(["Section", exam.section or ""])
        ws2.append([])
        ws2.append(["Subjects"])
        ws2.append(["subject_id", "code", "name", "full_marks", "pass_marks"])

        subjects_qs = Subject.objects.filter(school_class_id=exam.class_name_id)
        if getattr(exam, "subject_id", None):
            subjects_qs = subjects_qs.filter(id=exam.subject_id)
        if exam.class_name.sections:
            subjects_qs = subjects_qs.filter(section=(exam.section or ""))
        else:
            subjects_qs = subjects_qs.filter(section="")
        for subj in subjects_qs.order_by("code", "id"):
            ws2.append([subj.id, subj.code, subj.name, subj.full_marks, subj.pass_marks])

        ws2.append([])
        ws2.append(["Students"])
        ws2.append(["student_id", "name"])
        students_qs = Student.objects.filter(school_class_id=exam.class_name_id)
        if exam.class_name.sections:
            students_qs = students_qs.filter(section=(exam.section or ""))
        else:
            students_qs = students_qs.filter(section="")
        for st in students_qs.order_by("first_name", "last_name", "id")[:50]:
            ws2.append([st.id, f"{st.first_name} {st.last_name}".strip()])

        buff = BytesIO()
        wb.save(buff)
        buff.seek(0)
        filename = f"marks_upload_exam_{exam.id}.xlsx"
        return FileResponse(
            buff,
            as_attachment=True,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @action(detail=False, methods=["post"], url_path="import-excel")
    def import_excel(self, request):
        exam_id = request.data.get("exam")
        f = request.FILES.get("file")
        if not exam_id:
            raise ValidationError({"exam": "exam is required."})
        if not f:
            raise ValidationError({"file": "file is required."})

        try:
            exam = Exam.objects.select_related("class_name").get(id=exam_id)
        except Exam.DoesNotExist:
            raise ValidationError({"exam": "Exam not found."})
        if exam.status == Exam.STATUS_PUBLISHED:
            raise ValidationError({"detail": "Cannot modify marks after results are published."})

        if getattr(request.user, "role", None) == "TEACHER":
            ok = ClassTeacher.objects.filter(
                school_class_id=exam.class_name_id, section=(exam.section or ""), teacher__user_id=request.user.id
            ).exists()
            if not ok:
                raise PermissionDenied("Not your assigned class.")

        try:
            import openpyxl
        except Exception as e:  # pragma: no cover
            raise ValidationError({"detail": "openpyxl is required on the server for Excel import."}) from e

        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception as e:
            raise ValidationError({"file": f"Invalid Excel file: {e}"})

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValidationError({"file": "Excel is empty."})

        header = [str(c or "").strip().lower() for c in (rows[0] or [])]
        col_map = {name: idx for idx, name in enumerate(header) if name}
        required = ["student_id", "subject_id", "marks_obtained"]
        missing_cols = [c for c in required if c not in col_map]
        if missing_cols:
            raise ValidationError({"file": f"Missing required columns: {', '.join(missing_cols)}"})

        created = 0
        updated = 0
        errors = []

        for row_idx, row in enumerate(rows[1:], start=2):
            sid = row[col_map["student_id"]] if col_map.get("student_id") is not None else None
            subid = row[col_map["subject_id"]] if col_map.get("subject_id") is not None else None
            marks = row[col_map["marks_obtained"]] if col_map.get("marks_obtained") is not None else None
            remarks = ""
            if "remarks" in col_map:
                remarks = str(row[col_map["remarks"]] or "").strip()

            if sid in (None, "") and subid in (None, ""):
                continue
            try:
                sid = int(sid)
                subid = int(subid)
                if getattr(exam, "subject_id", None) and subid != exam.subject_id:
                    raise ValueError(f"Only subject_id={exam.subject_id} is allowed for this exam.")
                marks_val = None if marks in (None, "") else marks
                obj, was_created = StudentExamMark.objects.update_or_create(
                    student_id=sid,
                    exam=exam,
                    subject_id=subid,
                    defaults={"marks_obtained": marks_val, "remarks": remarks},
                )
                created += 1 if was_created else 0
                updated += 0 if was_created else 1
            except Exception as e:
                errors.append({"row": row_idx, "student_id": sid, "subject_id": subid, "error": str(e)})

        create_audit_log(
            action_type=AuditLog.ACTION_MARKS_BULK_UPLOAD,
            user=request.user,
            details={"exam": exam.id, "created": created, "updated": updated, "errors": len(errors), "source": "excel"},
        )
        return Response({"exam": exam.id, "created": created, "updated": updated, "errors": errors})


class ResultViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Result.objects.select_related("student", "exam", "exam__class_name").all()
    serializer_class = ResultSerializer
    rbac_path = "/portal/exam/results"

    def get_permissions(self):
        self.permission_classes = [permissions.IsAuthenticated, HasPortalPermission, IsAdmin | IsTeacher | IsStudent]
        return super().get_permissions()

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        exam_id = self.request.query_params.get("exam")
        student_id = self.request.query_params.get("student")
        q = (self.request.query_params.get("q") or "").strip()
        if exam_id:
            qs = qs.filter(exam_id=exam_id)
        if student_id:
            qs = qs.filter(student_id=student_id)
        if q:
            if q.isdigit():
                qs = qs.filter(student_id=int(q))
            else:
                qs = qs.filter(
                    Q(student__first_name__icontains=q)
                    | Q(student__last_name__icontains=q)
                    | Q(student__user__username__icontains=q)
                )

        role = getattr(user, "role", None)
        if role == "STUDENT":
            return qs.filter(student__user_id=user.id, published_status=Result.STATUS_PUBLISHED)
        if role == "TEACHER":
            # Teachers can see results for their assigned classrooms (generated/published).
            assignments = list(
                ClassTeacher.objects.filter(teacher__user_id=user.id).values_list("school_class_id", "section")
            )
            if not assignments:
                return qs.none()
            clause = Q()
            for class_id, section in assignments:
                clause |= Q(exam__class_name_id=class_id, exam__section=(section or ""))
            return qs.filter(clause)
        return qs

    @action(detail=True, methods=["get"], url_path="pdf")
    def pdf(self, request, pk=None):
        result = self.get_object()
        user = request.user
        role = getattr(user, "role", None)

        if role == "STUDENT" and getattr(result.student, "user_id", None) != user.id:
            raise PermissionDenied("Not allowed.")
        if role == "TEACHER" and not IsAdminOrTeacherWithClass().has_object_permission(request, self, result.exam):
            raise PermissionDenied("Not allowed.")

        content, filename = build_report_card_pdf(result=result)
        return FileResponse(BytesIO(content), as_attachment=True, filename=filename, content_type="application/pdf")


class RankingViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Merit ranking list for an exam (descending total marks).
    """

    queryset = Result.objects.select_related("student", "exam", "exam__class_name").all()
    serializer_class = ResultSerializer
    rbac_path = "/portal/exam/rankings"

    def get_permissions(self):
        self.permission_classes = [permissions.IsAuthenticated, HasPortalPermission, IsAdmin | IsTeacher]
        return super().get_permissions()

    def get_queryset(self):
        qs = super().get_queryset()
        exam_id = self.request.query_params.get("exam")
        if exam_id:
            qs = qs.filter(exam_id=exam_id)
        return qs.order_by("rank", "-total_marks", "student_id")


class PromotionViewSet(viewsets.ModelViewSet):
    queryset = Promotion.objects.select_related("student", "from_class", "to_class", "promoted_by").all()
    serializer_class = PromotionSerializer
    rbac_path = "/portal/exam/promotions"
    rbac_action_map = {"bulk": "create"}

    def get_permissions(self):
        if self.action in {"create", "update", "partial_update", "destroy", "bulk"}:
            self.permission_classes = [permissions.IsAuthenticated, HasPortalPermission, IsAdmin]
        else:
            self.permission_classes = [permissions.IsAuthenticated, HasPortalPermission]
        return super().get_permissions()

    @action(detail=False, methods=["post"], url_path="bulk")
    def bulk(self, request):
        academic_year = (request.data.get("academic_year") or "").strip()
        from_class = request.data.get("from_class")
        to_class = request.data.get("to_class")
        student_ids = request.data.get("student_ids") or []
        from_section = (request.data.get("from_section") or "").strip().upper()
        to_section = (request.data.get("to_section") or "").strip().upper()
        exam_id = request.data.get("exam", None)

        if not academic_year:
            raise ValidationError({"academic_year": "academic_year is required."})
        if not from_class or not to_class:
            raise ValidationError({"detail": "from_class and to_class are required."})
        if not isinstance(student_ids, list) or not student_ids:
            raise ValidationError({"student_ids": "student_ids must be a non-empty list."})

        students = list(Student.objects.filter(id__in=student_ids, school_class_id=from_class))
        if len(students) != len(set(student_ids)):
            raise ValidationError({"student_ids": "Some students not found in from_class."})
        if from_section:
            bad = [s.id for s in students if (s.section or "").strip().upper() != from_section]
            if bad:
                raise ValidationError({"student_ids": f"Some students not in section {from_section}: {bad}"})

        exam = None
        if exam_id:
            try:
                exam = Exam.objects.get(id=exam_id)
            except Exam.DoesNotExist:
                raise ValidationError({"exam": "Exam not found."})

        created = 0
        for s in students:
            Promotion.objects.create(
                student=s,
                from_class_id=from_class,
                from_section=from_section,
                to_class_id=to_class,
                to_section=to_section,
                promoted_by=request.user,
                academic_year=academic_year,
                exam=exam,
            )
            s.school_class_id = to_class
            s.section = to_section
            s.save(update_fields=["school_class", "section"])
            created += 1

        create_audit_log(
            action_type=AuditLog.ACTION_PROMOTION_BULK,
            user=request.user,
            details={"academic_year": academic_year, "from_class": from_class, "to_class": to_class, "students": created},
        )
        return Response({"created": created})


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.select_related("user").all()
    serializer_class = AuditLogSerializer
    rbac_path = "/portal/exam/audit-logs"

    def get_permissions(self):
        self.permission_classes = [permissions.IsAuthenticated, HasPortalPermission, IsAdmin]
        return super().get_permissions()
